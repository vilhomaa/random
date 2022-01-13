# A program for tracking Wow Auction House price data 

# The program extracts pricedata from aux addon's savedvariables file and saves them into a xlsx file for calculating profits from farming primal life
# The program first finds the trackable items from your xlsx file, then exctracts their id:s from wow_items.json and then finds the prices from aux-addon.lua file in your wtf folder
# Finally, the program writes the information if there are dates written on row 2 but no price data in rows 28-38
# An example xlsx file "Bog.xslx" is included. 

import json
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import re

# These have to be changed depending on your own game file paths and 
os.chdir("Z:\peleej\World of Warcraft tbc")
path = "muistiinpanot/bog.xlsx"


def get_items_to_track(path): 

    wb = load_workbook(path)
    ws = wb["Sheet1"]

    temp = [i[0].value.lower().strip() for i in ws.iter_rows(min_col = 1, max_col = 1, min_row = 13) if i[0].value is not None and i[0].value != "Prices" and i[0].value != "Drops"]

    items = list(set(temp))
    return items


def get_item_ids(path):
    tracked_items = get_items_to_track(path)

    with open("muistiinpanot/wow_items.json") as json_file:
        data = json.load(json_file)

    items_dict = {}

    for juttu in tracked_items:
        items_dict[juttu] = data["aux_item_ids"][juttu]

    return items_dict



def get_lua_data(path):
    items = get_item_ids(path)


    f = open("Z:/peleej/World of Warcraft tbc/WTF/Account/LADEOWNAA/SavedVariables/aux-addon.lua", 'r')
    lua_file = f.readlines()
    f.close()


    start = [i for i in lua_file if "history" in i][0]
    end = [i for i in lua_file if "post" in i][1]

    lua_file.index(start)
    lua_file.index(end)

    lua_file = lua_file[lua_file.index(start):lua_file.index(end)]


    for i in items:

        items[i] = [items[i]]

        items[i].append([j for j in lua_file if str(items[i][0])+':' in j][0])
    
    return items



def get_price_to_write(items_dict):

    for i in items_dict:
        exp = items_dict[i][1]
        items_dict[i][1] = (re.findall('\d+', re.findall('#+[0-9]{3,}',exp)[0]))[0]
    return items_dict



def write(path, items_dict):

    wb = load_workbook(path)
    ws = wb["Sheet1"]
    

    wb = load_workbook(path)
    ws = wb["Sheet1"]

    dates_written = len([i[0].value for i in ws.iter_cols(min_row = 2, max_row = 2) if i[0].value is not None])
    values_written = len([i[0].value for i in ws.iter_cols(min_row = 28, max_row = 28) if i[0].value is not None])


    diff_cols = dates_written-values_written

    if diff_cols >0:

        for row in ws.iter_rows(min_row = 28, max_row = 38, ):# min_col = dates_written+(-diff_cols), max_col = dates_written
            
            for item in items_dict:
                
                if row[0].value.lower().strip() == item:

                    for i in row[(dates_written-diff_cols):dates_written]:
                        i.value = round(int(items_dict[item][1])/10000,2)



    wb.save(path)


items_dict = get_lua_data(path)
items_dict = get_price_to_write(items_dict)
write(path,items_dict)

print("Prices updated!")
